"""
Student Result Verification System - Flask Backend
Manages student data lookup, verification actions, and audit trail logging
"""

from flask import Flask, render_template, request, jsonify, send_from_directory
from flask_cors import CORS
import csv
import json
import os
import tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

# Configuration
BASE_DIR = Path(__file__).parent
# Use BASE_DIR for CSV file too (assumes CSV is in same folder as app.py)
CSV_FILE = BASE_DIR / "AI_Data.csv"
EXCEL_FILE = BASE_DIR / "AI_Data_Lookup_Tracker.xlsx"
AUDIT_TRAIL_FILE = BASE_DIR / "audit_trail.json"
LOOKUP_TRACKING_FILE = BASE_DIR / "lookup_history.json"

def load_students_from_excel():
    """Load student records from the existing Excel tracker workbook."""
    students = {}
    try:
        if not EXCEL_FILE.exists():
            return students
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not header_row or 'Roll No' not in header_row:
            return students

        roll_index = header_row.index('Roll No')
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            roll_no = row[roll_index]
            if roll_no:
                student = {header_row[i]: row[i] for i in range(len(header_row))}
                students[str(roll_no).strip().upper()] = student
        return students
    except Exception as e:
        print(f"Error loading Excel data: {e}")
        return {}


# Load students from CSV, with Excel merge to support partial CSV datasets
def load_students():
    """Load student records from CSV file, merging with Excel data if available."""
    students = {}
    try:
        csv_count = 0
        if CSV_FILE.exists():
            with open(CSV_FILE, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    if row and row.get('Roll No'):
                        roll_no = str(row['Roll No']).strip().upper()
                        students[roll_no] = row
                        csv_count += 1

        if EXCEL_FILE.exists():
            excel_students = load_students_from_excel()
            if excel_students:
                # Merge Excel data so missing records are available
                for roll_no, student in excel_students.items():
                    if roll_no not in students:
                        students[roll_no] = student
                if len(excel_students) > csv_count:
                    print(f"Loaded {len(excel_students)} students from Excel; CSV had {csv_count} records.")
        if not students:
            print("No student data found in CSV or Excel.")
        return students
    except Exception as e:
        print(f"Error loading CSV: {e}")
        if EXCEL_FILE.exists():
            print("Falling back to Excel data source.")
            return load_students_from_excel()
        return {}

# Load students at startup
STUDENTS = load_students()

def create_excel_from_csv():
    """Create/recreate Excel file from CSV with formatting"""
    try:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Records"
        
        # Read CSV and write to Excel
        with open(CSV_FILE, 'r', encoding='utf-8') as f:
            csv_reader = csv.reader(f)
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format header row
                    if row_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF", size=11)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Save workbook
        if not save_workbook_safe(wb, EXCEL_FILE):
            return False
        print(f"Excel file created: {EXCEL_FILE}")
        return True
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False


def save_workbook_safe(wb, file_path):
    """Save workbook safely, with a temporary fallback if possible."""
    try:
        wb.save(file_path)
        return True
    except PermissionError as pe:
        print(f"Permission denied saving {file_path}: {pe}")
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=file_path.parent) as tmp:
                temp_path = Path(tmp.name)
            wb.save(temp_path)
            os.replace(temp_path, file_path)
            return True
        except Exception as e:
            print(f"Safe save fallback failed: {e}")
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception:
                pass
            return False
    except Exception as e:
        print(f"Unexpected error saving workbook: {e}")
        return False


def highlight_student_in_excel(roll_no, student_name):
    """Highlight student row in Excel file with red background and timestamp"""
    try:
        # Create Excel if it doesn't exist
        if not EXCEL_FILE.exists():
            create_excel_from_csv()
        
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Get header row to find Roll No column (do this once)
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        if 'Roll No' not in header_row:
            print(f"Roll No column not found in Excel header")
            return False
        
        roll_col = header_row.index('Roll No') + 1
        print(f"DEBUG: Looking for roll_no='{roll_no}' in column {roll_col}")
        
        # Find and highlight the student row
        found = False
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        
        for row_idx in range(2, ws.max_row + 1):  # Skip header, iterate through all rows
            roll_cell = ws.cell(row=row_idx, column=roll_col)
            roll_cell_value = str(roll_cell.value).strip() if roll_cell.value else ""
            
            # Debug: print first few rows
            if row_idx <= 5:
                print(f"DEBUG: Row {row_idx}: roll_cell_value='{roll_cell_value}'")
            
            # If this is the student we're looking for
            if roll_cell_value == roll_no.strip():
                print(f"DEBUG: FOUND MATCH at row {row_idx}!")
                # Highlight entire row - iterate through ALL columns
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = red_fill
                    cell.font = white_font
                
                # Add timestamp in a new column
                last_col = ws.max_column + 1
                timestamp_cell = ws.cell(row=row_idx, column=last_col, value=datetime.now().isoformat())
                timestamp_cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                timestamp_cell.font = Font(color="FFFFFF", size=9)
                
                found = True
                break
        
        if found:
            if not save_workbook_safe(wb, EXCEL_FILE):
                print(f"Error: could not save highlighted Excel file for {roll_no}")
                wb.close()
                return False
            print(f"✓ Highlighted student {roll_no} ({student_name}) in Excel")
            return True
        else:
            print(f"DEBUG: Student {roll_no} not found in Excel!")
            wb.close()
            return False
            
    except Exception as e:
        print(f"Error highlighting student in Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def log_lookup(roll_no, student_name):
    """Log student lookup with timestamp"""
    try:
        lookup_history = []
        if LOOKUP_TRACKING_FILE.exists():
            with open(LOOKUP_TRACKING_FILE, 'r', encoding='utf-8') as f:
                lookup_history = json.load(f)
        
        entry = {
            "timestamp": datetime.now().isoformat(),
            "roll_no": roll_no,
            "student_name": student_name
        }
        lookup_history.append(entry)
        
        with open(LOOKUP_TRACKING_FILE, 'w', encoding='utf-8') as f:
            json.dump(lookup_history, f, indent=2, ensure_ascii=False)
        
        return True
    except Exception as e:
        print(f"Error logging lookup: {e}")
        return False

def log_audit_trail(officer_name, roll_no, student_name, action, comments=""):
    """Log verification action to audit trail"""
    try:
        # Load existing audit trail
        audit_trail = []
        if AUDIT_TRAIL_FILE.exists():
            with open(AUDIT_TRAIL_FILE, 'r', encoding='utf-8') as f:
                audit_trail = json.load(f)
        
        # Add new entry
        entry = {
            "timestamp": datetime.now().isoformat(),
            "officer_name": officer_name,
            "roll_no": roll_no,
            "student_name": student_name,
            "action": action,
            "comments": comments
        }
        audit_trail.append(entry)
        
        # Save audit trail
        with open(AUDIT_TRAIL_FILE, 'w', encoding='utf-8') as f:
            json.dump(audit_trail, f, indent=2, ensure_ascii=False)
        
        return True
    except Exception as e:
        print(f"Error logging audit trail: {e}")
        return False

@app.route('/')
def index():
    """Serve the main page"""
    return render_template('index.html')

@app.route('/api/student/<roll_no>')
def get_student(roll_no):
    """Get student details by roll number"""
    roll_no = roll_no.strip().upper()
    
    # Search for student (case-insensitive)
    student = None
    for key, value in STUDENTS.items():
        if key.strip().upper() == roll_no:
            student = value
            break
    
    if student:
        # Log the lookup (but don't highlight yet)
        log_lookup(roll_no, student.get('Name', 'Unknown'))
        
        return jsonify({
            "success": True,
            "data": student
        })
    else:
        return jsonify({
            "success": False,
            "message": "Student not found"
        }), 404

@app.route('/api/verify', methods=['POST'])
def verify_student():
    """Submit verification action"""
    data = request.json
    
    try:
        officer_name = data.get('officer_name', '').strip()
        roll_no = data.get('roll_no', '').strip()
        action = data.get('action', '').strip()
        comments = data.get('comments', '').strip()
        student_name = data.get('student_name', '').strip()
        
        if not all([officer_name, roll_no, action, student_name]):
            return jsonify({
                "success": False,
                "message": "Missing required fields"
            }), 400
        
        # Validate action
        valid_actions = ['Issued', 'Corrected']
        if action not in valid_actions:
            return jsonify({
                "success": False,
                "message": f"Invalid action. Must be one of: {', '.join(valid_actions)}"
            }), 400
        
        # Log to audit trail
        if log_audit_trail(officer_name, roll_no, student_name, action, comments):
            highlight_error = None
            # Only highlight in Excel if "Issued" is selected
            if action == 'Issued':
                if not highlight_student_in_excel(roll_no, student_name):
                    highlight_error = "Excel file could not be updated. Close the spreadsheet if it is open and retry."
            
            response = {
                "success": True,
                "message": f"Verification recorded: {action}",
                "timestamp": datetime.now().isoformat()
            }
            if highlight_error:
                response["warning"] = highlight_error
            return jsonify(response)
        else:
            return jsonify({
                "success": False,
                "message": "Failed to save verification"
            }), 500
            
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"
        }), 500

@app.route('/api/audit-trail')
def get_audit_trail():
    """Get audit trail records (admin view)"""
    try:
        if AUDIT_TRAIL_FILE.exists():
            with open(AUDIT_TRAIL_FILE, 'r', encoding='utf-8') as f:
                audit_trail = json.load(f)
            return jsonify({
                "success": True,
                "records": audit_trail
            })
        else:
            return jsonify({
                "success": True,
                "records": []
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error loading audit trail: {str(e)}"
        }), 500

@app.route('/api/stats')
def get_stats():
    """Get verification statistics"""
    try:
        if AUDIT_TRAIL_FILE.exists():
            with open(AUDIT_TRAIL_FILE, 'r', encoding='utf-8') as f:
                audit_trail = json.load(f)
            
            stats = {
                "total_verifications": len(audit_trail),
                "issued": sum(1 for r in audit_trail if r['action'] == 'Issued'),
                "corrected": sum(1 for r in audit_trail if r['action'] == 'Corrected')
            }
            return jsonify({
                "success": True,
                "stats": stats
            })
        else:
            return jsonify({
                "success": True,
                "stats": {
                    "total_verifications": 0,
                    "issued": 0,
                    "corrected": 0
                }
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"
        }), 500

@app.errorhandler(404)
def handle_404(error):
    if request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": "API endpoint not found"
        }), 404
    return error, 404

@app.errorhandler(500)
def handle_500(error):
    if request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "message": "Server error occurred"
        }), 500
    return error, 500

@app.route('/api/lookup-history')
def get_lookup_history():
    """Get lookup history (students searched)"""
    try:
        if LOOKUP_TRACKING_FILE.exists():
            with open(LOOKUP_TRACKING_FILE, 'r', encoding='utf-8') as f:
                lookup_history = json.load(f)
            return jsonify({
                "success": True,
                "records": lookup_history
            })
        else:
            return jsonify({
                "success": True,
                "records": []
            })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error loading lookup history: {str(e)}"
        }), 500

@app.route('/api/download-excel')
def download_excel():
    """Provide Excel file for download"""
    try:
        if EXCEL_FILE.exists():
            return send_from_directory(
                str(EXCEL_FILE.parent),
                EXCEL_FILE.name,
                as_attachment=True,
                download_name=f"Student_Records_Lookup_Tracker.xlsx"
            )
        else:
            return jsonify({
                "success": False,
                "message": "Excel file not found"
            }), 404
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"
        }), 500

# ============================================================
# NEW ENDPOINT - Check if student has been issued a certificate
# ============================================================
@app.route('/api/student-issued-status/<roll_no>')
def get_student_issued_status(roll_no):
    """Check if a student has been issued a certificate"""
    try:
        if AUDIT_TRAIL_FILE.exists():
            with open(AUDIT_TRAIL_FILE, 'r', encoding='utf-8') as f:
                audit_trail = json.load(f)
            
            # Find all issued records for this student
            issued_records = [r for r in audit_trail if r['roll_no'] == roll_no and r['action'] == 'Issued']
            
            if issued_records:
                # Get the most recent issued record
                latest = issued_records[-1]
                return jsonify({
                    "success": True,
                    "issued": True,
                    "timestamp": latest['timestamp'],
                    "officer_name": latest['officer_name'],
                    "comments": latest.get('comments', '')
                })
        
        return jsonify({"success": True, "issued": False})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

if __name__ == '__main__':
    print(f"Loaded {len(STUDENTS)} students from CSV")
    print(f"CSV file: {CSV_FILE}")
    print(f"Excel tracking file: {EXCEL_FILE}")
    print(f"Audit trail: {AUDIT_TRAIL_FILE}")
    
    # Create Excel file on startup
    print("\nCreating Excel tracking file...")
    if create_excel_from_csv():
        print("✓ Excel file ready for lookup tracking")
    
    app.run(host='0.0.0.0', port=5000)