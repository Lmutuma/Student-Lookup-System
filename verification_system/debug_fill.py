from openpyxl import load_workbook

wb = load_workbook('AI_Data_Lookup_Tracker.xlsx')
ws = wb.active

# Get headers
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
roll_col_idx = header_row.index('Roll No') + 1 if 'Roll No' in header_row else None
name_col_idx = header_row.index('Name') + 1 if 'Name' in header_row else 2

print('=== DETAILED FILL CHECK ===')
print()

# Check first 20 rows
for row_idx in range(2, min(20, ws.max_row + 1)):
    first_cell = ws.cell(row=row_idx, column=1)
    name_cell = ws.cell(row=row_idx, column=name_col_idx)
    roll_cell = ws.cell(row=row_idx, column=roll_col_idx) if roll_col_idx else None
    
    # Get all fill info
    if first_cell.fill:
        print(f'Row {row_idx}: {name_cell.value} ({roll_cell.value if roll_cell else "N/A"})')
        print(f'  Fill object: {first_cell.fill}')
        print(f'  Start color: {first_cell.fill.start_color}')
        if first_cell.fill.start_color:
            print(f'    RGB: {first_cell.fill.start_color.rgb}')
            print(f'    Index: {first_cell.fill.start_color.index}')
        print()
