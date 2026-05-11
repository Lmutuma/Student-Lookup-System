from openpyxl import load_workbook

wb = load_workbook('AI_Data_Lookup_Tracker.xlsx')
ws = wb.active

# Check headers first
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print(f'Headers: {header_row}')
print()

# Find Roll No column
roll_col_idx = None
if 'Roll No' in header_row:
    roll_col_idx = header_row.index('Roll No') + 1
    print(f'Roll No is at column {roll_col_idx}')
else:
    print('Roll No column not found!')
print()

# Check a few rows to see if they have highlighting
print('Checking for highlighted rows (first 15 data rows):')
print()

highlighted_rows = []
for row_idx in range(2, min(20, ws.max_row + 1)):
    first_cell = ws.cell(row=row_idx, column=1)
    roll_cell = ws.cell(row=row_idx, column=roll_col_idx) if roll_col_idx else None
    
    # Check if cell has red fill
    if first_cell.fill and first_cell.fill.start_color:
        try:
            fill_rgb = first_cell.fill.start_color.rgb
            if fill_rgb and 'FF' in str(fill_rgb):  # Red-ish color
                name_value = ws.cell(row=row_idx, column=1).value
                roll_value = roll_cell.value if roll_cell else 'N/A'
                print(f'Row {row_idx}: {name_value} ({roll_value}) - Fill: {fill_rgb}')
                highlighted_rows.append(row_idx)
        except:
            pass

if not highlighted_rows:
    print('No highlighted rows found!')
print()
print(f'Total highlighted rows: {len(highlighted_rows)}')
