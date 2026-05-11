from openpyxl import load_workbook

wb = load_workbook('AI_Data_Lookup_Tracker.xlsx')
ws = wb.active

# Get headers
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
roll_col_idx = header_row.index('Roll No') + 1 if 'Roll No' in header_row else None
name_col_idx = header_row.index('Name') + 1 if 'Name' in header_row else 2

print('=== HIGHLIGHTED STUDENTS (Issued Certificates) ===')
print()

highlighted_rows = []
for row_idx in range(2, ws.max_row + 1):
    first_cell = ws.cell(row=row_idx, column=1)
    
    # Check if cell has red fill (00FF0000 or similar)
    if first_cell.fill and first_cell.fill.start_color:
        try:
            fill_rgb = first_cell.fill.start_color.rgb
            if fill_rgb and 'FF' in str(fill_rgb):  # Red color indicator
                name_value = ws.cell(row=row_idx, column=name_col_idx).value
                roll_value = ws.cell(row=row_idx, column=roll_col_idx).value if roll_col_idx else 'N/A'
                
                # Get timestamp from last column
                last_col = ws.max_column
                timestamp_value = ws.cell(row=row_idx, column=last_col).value
                
                print(f'✓ {name_value} ({roll_value})')
                print(f'  Timestamp: {timestamp_value}')
                print()
                highlighted_rows.append((name_value, roll_value))
        except:
            pass

if not highlighted_rows:
    print('No highlighted rows found!')
else:
    print(f'Total Issued Certificates: {len(highlighted_rows)}')
